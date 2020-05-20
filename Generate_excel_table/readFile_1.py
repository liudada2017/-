# -*- coding:gb2312 -*-
import os
import re
import openpyxl
from openpyxl.styles import Font
#��ȡLog�ļ�������ȡ�ļ��еĹؼ���
def read_log_data(filename):
    with open(filename, 'r') as file:
        frame_count_list = []#���֡�;����б�
        time_list  = [] #���ʱ���б�
        first_win = True #�Ƿ�洢win��ʱ��
        save_frame_count = False#�Ƿ�洢֡�;���

        while True:
            lines = file.readline()  # ���ж�ȡ����
            if not lines:
                break
            item_win = [i for i in lines.split(']')]
            item_over = [i for i in lines.split(':')]

            if len(item_over)!=1 and item_over[-1]!='\n':
                if item_over[-1].split(',')[-1].split()[-1].find('win')!=-1:
                    if first_win:
                        time_list.append(item_win[0].split(',')[0].split()[-1])
                        first_win = False
                        save_frame_count = True

                elif item_over[len(item_over)-2].find('episode over')!= -1:
                   if save_frame_count:
                       index = item_over[-1].split(',')[0].split('=')[-1]
                       count = item_over[-1].split(',')[1].split('=')[-1]
                       frame_count_list.append([index,count])
                       first_win = False
                       save_frame_count = False

                elif item_over[-1].find('Epsiode over')!=-1:
                    first_win = False
                else:
                    first_win = True
    file.close()
    for i in range(0,len(time_list)):
        frame_count_list[i].insert(0,time_list[i])
    return frame_count_list

#����������д����csv�ļ���
def write_to_csv(fileList, dir_path):
    n = 0 #������λ��
    wb = openpyxl.Workbook()#�½�һ��������
    for file_name in fileList:
        sheet_name = file_name.split('\\')[-1].split('.')[0]#��ȡ�������
        ws = wb.create_sheet(title= sheet_name, index=n) #����һ�����
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        ws['C1'].font = Font(bold=True)
        ws['D1'].font = Font(bold=True)
        ws['A1'].value = 'ʤ������'
        ws['B1'].value = 'ʱ��'
        ws['C1'].value = '֡��'
        ws['D1'].value = '����'
        sum = read_log_data(file_name)
        for i in range(1,len(sum)+1):#����Ԫ������д����
            ws.cell(i+1, 1).value = i
            for j in range(1,4):
                ws.cell(i+1, j+1).value = sum[i-1][j-1]
        n = n+1
    new_count_table(n, wb)
    wb.save(dir_path + '\\' + 'result.xls')#����excel�ļ�
    print('����ɹ�,�����ַΪ��',dir_path + '\\' + 'result.xls')

#�½����table,����д������
def new_count_table(n, wb):
    ws = wb.create_sheet(title='Count', index=n)
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)
    ws['D1'].font = Font(bold=True)
    ws['E1'].font = Font(bold=True)
    ws['F1'].font = Font(bold=True)
    ws['G1'].font = Font(bold=True)
    ws['A1'].value = 'id'
    ws['B1'].value = 'ʱ��'
    ws['C1'].value = '֡��'
    ws['D1'].value = '����'
    ws['E1'].value = '�ֻ����'
    ws['F1'].value = 'Ӯ��ʱ��'
    ws['G1'].value = '��ʼ����ʱ��'


#��ȡ�ļ����µ�������־�ļ�
def get_files(dir_path):
    fileList = []#���������־�ļ�
    files = os.listdir(dir_path)#��ȡ�����ļ�
    ptn = re.compile('.*\.log') #log��β���ļ�������ʽ
    for f in files: #���������ļ�����ȡlog��β����־�ļ�
        if (os.path.isfile(dir_path + '\\' + f)):
            res = ptn.match(f)
            if (res != None) and f != 'train.log':
                fileList.append(dir_path + '\\' + res.group())
        else:
            print('��Ч�ļ�')
    return fileList


if __name__ == "__main__":
    dir_path = 'F:\\New_Game\\QQ�ɳ�\\log_2020.5.13_1-4_6' #�����־���ļ���
    fileList = get_files(dir_path)
    write_to_csv(fileList, dir_path)
