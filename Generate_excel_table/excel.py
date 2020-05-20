from openpyxl.styles import Font
import openpyxl

wb = openpyxl.Workbook()

for i in range(4):
    ws = wb.create_sheet(title = 'agent_' + str(i),index = i)
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)
    ws['D1'].font = Font(bold=True)
    ws['A1'].value = '胜利次数'
    ws['B1'].value = '时间'
    ws['C1'].value = '帧号'
    ws['D1'].value = '局数'
wb.save('test.xlsx')